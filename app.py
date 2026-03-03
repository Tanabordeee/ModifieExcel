import streamlit as st
from openpyxl import load_workbook
from datetime import datetime, timedelta
import random
import tempfile

st.title("🎲 Random DateTime Excel Tool")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file:
    st.info("Sheet: แรกสุดของไฟล์ | Column: A (ประทับเวลา)")

    st.subheader("เลือกช่วงวันเวลา")

    col1, col2 = st.columns(2)

    with col1:
        start_date = st.date_input("Start Date")
        start_time = st.time_input("Start Time", value=datetime.min.time())

    with col2:
        end_date = st.date_input("End Date")
        end_time = st.time_input("End Time", value=datetime.max.time())

    start_dt = datetime.combine(start_date, start_time)
    end_dt = datetime.combine(end_date, end_time)

    if st.button("Random & Download"):
        wb = load_workbook(uploaded_file)
        ws = wb[wb.sheetnames[0]]  # sheet แรก

        # column A index = 0
        col_index = 0
        rows = []
        for row in ws.iter_rows(min_row=2):
            cell = row[col_index]
            rows.append(cell)

        total_seconds = int((end_dt - start_dt).total_seconds())

        # สุ่มวันเวลา
        random_dates = [
            start_dt + timedelta(seconds=random.randint(0, total_seconds))
            for _ in rows
        ]

        # เรียงวัน
        random_dates.sort()

        # ใส่กลับ excel
        for cell, new_dt in zip(rows, random_dates):
            original_format = cell.number_format
            cell.value = new_dt
            cell.number_format = original_format

        # สร้าง temp file ให้ user download
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            st.download_button(
                "Download File",
                data=open(tmp.name, "rb"),
                file_name="output.xlsx"
            )

        st.success(f"✅ Random วันเวลาเรียบร้อย จำนวนแถว {len(rows)}")